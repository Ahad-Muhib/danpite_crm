import csv
import io
from datetime import timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Q, Sum, Count, Avg, F
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from clients.models import Client

from .forms import (
    ActivityQuickForm, CommentForm, DealForm, FollowUpForm,
    LeadAssignmentForm, LeadContactForm,
)
from .models import Activity, Comment, Deal, FollowUp, LeadContact


def _log_activity(lead=None, deal=None, activity_type='note', title='', description='', user=None, meta=None):
    Activity.objects.create(
        lead=lead, deal=deal, activity_type=activity_type,
        title=title, description=description,
        created_by=user, meta=meta or {},
    )


# ── Leads ──────────────────────────────────────────────────────

@login_required
def lead_list(request):
    if request.method == 'POST' and 'bulk_action' in request.POST:
        selected_ids = request.POST.getlist('selected_leads')
        action = request.POST.get('bulk_action')
        selected = LeadContact.objects.filter(pk__in=selected_ids)

        if not selected_ids:
            messages.error(request, 'Select at least one lead contact first.')
            return redirect('lead_list')

        if action == 'convert':
            converted_count = 0
            for lead in selected.filter(is_converted=False):
                Client.objects.create(
                    name=lead.name, email=lead.email, phone=lead.phone,
                    company=lead.company, website=lead.website, address=lead.address,
                    account_manager=request.user, lead_contact=lead,
                )
                lead.is_converted = True
                lead.save(update_fields=['is_converted', 'updated_at'])
                _log_activity(lead=lead, activity_type='conversion', title='Converted to client', user=request.user)
                converted_count += 1
            messages.success(request, f'{converted_count} lead contact(s) converted to client.')
        elif action == 'delete':
            deleted_count = selected.count()
            selected.delete()
            messages.success(request, f'{deleted_count} lead contact(s) deleted.')
        else:
            messages.error(request, 'Choose a bulk action first.')
        return redirect('lead_list')

    q = request.GET.get('q', '')
    src = request.GET.get('source', '')
    contact_type = request.GET.get('type', '')
    qs = LeadContact.objects.all().order_by('-id')
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(email__icontains=q) | Q(phone__icontains=q) | Q(company__icontains=q))
    if src:
        qs = qs.filter(lead_source=src)
    if contact_type in ('lead', 'deal'):
        qs = qs.filter(contact_type=contact_type)
    source_options = LeadContact.objects.exclude(lead_source='').values_list('lead_source', flat=True).distinct().order_by('lead_source')
    return render(request, 'leads/leads.html', {
        'leads': qs, 'q': q, 'source': src,
        'contact_type': contact_type, 'source_options': source_options,
    })


def _lead_source_options():
    saved = LeadContact.objects.exclude(lead_source='').values_list('lead_source', flat=True).distinct()
    return sorted(set(LeadContact.SOURCE_SUGGESTIONS) | set(saved))


@login_required
def lead_create(request):
    form = LeadContactForm(request.POST or None)
    if form.is_valid():
        obj = form.save(commit=False)
        obj.added_by = request.user
        obj.save()
        _log_activity(lead=obj, activity_type='note', title='Lead created', user=request.user)
        messages.success(request, 'Lead contact added.')
        return redirect('lead_list')
    return render(request, 'leads/lead_form.html', {
        'form': form, 'action': 'Add', 'lead_source_options': _lead_source_options(),
    })


@login_required
def lead_edit(request, pk):
    obj = get_object_or_404(LeadContact, pk=pk)
    form = LeadContactForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.save()
        _log_activity(lead=obj, activity_type='note', title='Lead updated', user=request.user)
        messages.success(request, 'Lead updated.')
        return redirect('lead_list')
    return render(request, 'leads/lead_form.html', {
        'form': form, 'action': 'Edit', 'lead': obj, 'lead_source_options': _lead_source_options(),
    })


@login_required
def lead_detail(request, pk):
    lead = get_object_or_404(LeadContact, pk=pk)
    deals = lead.deals.all()
    followups = lead.followups.all().order_by('-created_at')[:10]
    activities = lead.activities.all()[:20]
    comments = lead.comments.all()
    today = timezone.now().date()
    comment_form = CommentForm()
    activity_form = ActivityQuickForm()
    assignment_form = LeadAssignmentForm(initial={'lead_owner': lead.lead_owner})
    return render(request, 'leads/lead_detail.html', {
        'lead': lead, 'deals': deals, 'followups': followups,
        'activities': activities, 'comments': comments, 'today': today,
        'comment_form': comment_form, 'activity_form': activity_form,
        'assignment_form': assignment_form,
    })


@login_required
def lead_delete(request, pk):
    get_object_or_404(LeadContact, pk=pk).delete()
    messages.success(request, 'Lead deleted.')
    return redirect('lead_list')


@login_required
def lead_toggle_active(request, pk):
    obj = get_object_or_404(LeadContact, pk=pk)
    if obj.is_converted:
        Client.objects.filter(lead_contact=obj).delete()
        obj.is_converted = False
        obj.save()
        _log_activity(lead=obj, activity_type='status_change', title='Reverted to active lead', user=request.user)
        messages.success(request, f'{obj.name} set to active. Client removed.')
    else:
        Client.objects.create(
            name=obj.name, email=obj.email, phone=obj.phone,
            company=obj.company, website=obj.website, address=obj.address,
            account_manager=request.user, lead_contact=obj,
        )
        obj.is_converted = True
        obj.save()
        _log_activity(lead=obj, activity_type='conversion', title='Converted to client', user=request.user)
        messages.success(request, f'{obj.name} converted to client.')
    return redirect('lead_list')


@login_required
def lead_convert(request, pk):
    lead = get_object_or_404(LeadContact, pk=pk)
    Client.objects.create(
        name=lead.name, email=lead.email, phone=lead.phone,
        company=lead.company, website=lead.website, address=lead.address,
        account_manager=request.user, lead_contact=lead,
    )
    lead.is_converted = True
    lead.save()
    _log_activity(lead=lead, activity_type='conversion', title='Converted to client', user=request.user)
    messages.success(request, f'{lead.name} converted to client.')
    return redirect('lead_list')


@login_required
@require_POST
def lead_assign(request, pk):
    lead = get_object_or_404(LeadContact, pk=pk)
    form = LeadAssignmentForm(request.POST)
    if form.is_valid():
        old_owner = lead.lead_owner
        new_owner = form.cleaned_data['lead_owner']
        lead.lead_owner = new_owner
        lead.save(update_fields=['lead_owner', 'updated_at'])
        owner_name = new_owner.get_full_name() or new_owner.username if new_owner else 'Unassigned'
        _log_activity(lead=lead, activity_type='assignment', title=f'Assigned to {owner_name}', user=request.user)
        messages.success(request, f'Lead assigned to {owner_name}.')
    return redirect('lead_detail', pk=pk)


# ── Comments (Feature 5) ──────────────────────────────────────

@login_required
@require_POST
def comment_add(request, pk):
    lead = get_object_or_404(LeadContact, pk=pk)
    form = CommentForm(request.POST)
    if form.is_valid():
        comment = form.save(commit=False)
        comment.lead = lead
        comment.created_by = request.user
        comment.save()
        _log_activity(lead=lead, activity_type='note', title=f'Comment added: {comment.body[:50]}', user=request.user)
    return redirect('lead_detail', pk=pk)


@login_required
@require_POST
def comment_delete(request, pk, comment_pk):
    lead = get_object_or_404(LeadContact, pk=pk)
    get_object_or_404(Comment, pk=comment_pk, lead=lead).delete()
    messages.success(request, 'Comment deleted.')
    return redirect('lead_detail', pk=pk)


# ── Activity Quick Log (Feature 8) ────────────────────────────

@login_required
@require_POST
def activity_quick_log(request, pk):
    lead = get_object_or_404(LeadContact, pk=pk)
    act_type = request.POST.get('activity_type', 'call')
    form = ActivityQuickForm(request.POST)
    if form.is_valid():
        title = form.cleaned_data['title']
        desc = form.cleaned_data['description']
        _log_activity(lead=lead, activity_type=act_type, title=title, description=desc, user=request.user)
        messages.success(request, f'{act_type.title()} logged: {title}')
    return redirect('lead_detail', pk=pk)


# ── Deals ──────────────────────────────────────────────────────

@login_required
def deal_list(request):
    today = timezone.now().date()
    qs = Deal.objects.all().order_by('-created_at')
    return render(request, 'leads/deals.html', {'deals': qs, 'today': today})


@login_required
def deal_kanban(request):
    pipeline = request.GET.get('pipeline', 'sales')
    deals_by_stage = []
    for stage_key, stage_label in Deal.STAGE:
        deals = list(Deal.objects.filter(pipeline=pipeline, stage=stage_key).select_related('lead_contact', 'deal_agent').order_by('-created_at'))
        deals_by_stage.append((stage_key, stage_label, deals))
    pipeline_stats = Deal.objects.filter(pipeline=pipeline).aggregate(
        total_value=Sum('value'),
        deal_count=Count('id'),
        avg_value=Avg('value'),
    )
    return render(request, 'leads/kanban.html', {
        'deals_by_stage': deals_by_stage,
        'current_pipeline': pipeline, 'pipelines': Deal.PIPELINE,
        'pipeline_stats': pipeline_stats,
    })


@login_required
@require_POST
def deal_update_stage(request, pk):
    deal = get_object_or_404(Deal, pk=pk)
    new_stage = request.POST.get('stage', '')
    valid_stages = [s[0] for s in Deal.STAGE]
    if new_stage not in valid_stages:
        messages.error(request, 'Invalid stage.')
        return redirect('deal_kanban')

    old_stage = deal.get_stage_display()
    deal.stage = new_stage

    if new_stage == 'lost':
        lost_reason = request.POST.get('lost_reason', '')
        deal.lost_reason = lost_reason
        deal.save(update_fields=['stage', 'lost_reason', 'updated_at'])
        _log_activity(deal=deal, lead=deal.lead_contact, activity_type='deal_update',
                      title=f'Deal moved to Lost', description=f'Reason: {lost_reason}',
                      user=request.user, meta={'from': old_stage, 'to': 'Lost', 'reason': lost_reason})
    else:
        deal.save(update_fields=['stage', 'updated_at'])
        _log_activity(deal=deal, lead=deal.lead_contact, activity_type='deal_update',
                      title=f'Deal moved: {old_stage} → {deal.get_stage_display()}',
                      user=request.user, meta={'from': old_stage, 'to': deal.get_stage_display()})

    # Auto-convert lead to client when deal is won
    if new_stage == 'won' and deal.lead_contact and not deal.lead_contact.is_converted:
        lead = deal.lead_contact
        Client.objects.create(
            name=lead.name, email=lead.email, phone=lead.phone,
            company=lead.company, website=lead.website, address=lead.address,
            account_manager=request.user, lead_contact=lead,
        )
        lead.is_converted = True
        lead.save(update_fields=['is_converted', 'updated_at'])
        _log_activity(lead=lead, activity_type='conversion',
                      title=f'Auto-converted: Deal "{deal.deal_name}" won',
                      user=request.user)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': True, 'stage': new_stage})
    return redirect('deal_kanban', pipeline=deal.pipeline)


@login_required
def deal_create(request):
    form = DealForm(request.POST or None)
    if form.is_valid():
        deal = form.save()
        _log_activity(deal=deal, lead=deal.lead_contact, activity_type='deal_update', title=f'Deal created: {deal.deal_name}', user=request.user)
        messages.success(request, 'Deal created.')
        return redirect('deal_list')
    return render(request, 'leads/deal_form.html', {'form': form, 'action': 'Create', 'leads': LeadContact.objects.all().order_by('name')})


@login_required
def deal_edit(request, pk):
    obj = get_object_or_404(Deal, pk=pk)
    form = DealForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.save()
        messages.success(request, 'Deal updated.')
        return redirect('deal_list')
    return render(request, 'leads/deal_form.html', {'form': form, 'action': 'Edit', 'leads': LeadContact.objects.all().order_by('name')})


@login_required
def deal_delete(request, pk):
    get_object_or_404(Deal, pk=pk).delete()
    messages.success(request, 'Deal deleted.')
    return redirect('deal_list')


# ── Follow-ups (with Recurring, Feature 10) ───────────────────

@login_required
def followup_list(request):
    q = request.GET.get('q', '')
    ftype = request.GET.get('type', '')
    upcoming = request.GET.get('upcoming', '')
    today = timezone.now().date()
    qs = FollowUp.objects.select_related('lead', 'deal', 'created_by').all()
    if q:
        qs = qs.filter(Q(subject__icontains=q) | Q(lead__name__icontains=q) | Q(deal__deal_name__icontains=q) | Q(notes__icontains=q))
    if ftype in ('call', 'email', 'meeting', 'note', 'other'):
        qs = qs.filter(followup_type=ftype)
    if upcoming == '1':
        qs = qs.filter(next_followup_date__gte=today).order_by('next_followup_date')
    elif upcoming == '0':
        qs = qs.filter(next_followup_date__lt=today).order_by('-next_followup_date')
    else:
        qs = qs.order_by('-created_at')
    return render(request, 'leads/followups.html', {
        'followups': qs, 'q': q, 'ftype': ftype,
        'upcoming': upcoming, 'today': today,
    })


@login_required
def followup_create(request):
    initial = {}
    lead_id = request.GET.get('lead')
    deal_id = request.GET.get('deal')
    if lead_id:
        initial['lead'] = lead_id
    if deal_id:
        initial['deal'] = deal_id
    form = FollowUpForm(request.POST or None, initial=initial)
    if form.is_valid():
        obj = form.save(commit=False)
        obj.created_by = request.user
        obj.save()
        _log_activity(lead=obj.lead, deal=obj.deal, activity_type='followup',
                      title=f'Follow-up: {obj.subject}', user=request.user)
        if obj.is_recurring and obj.next_followup_date:
            _create_recurring_followup(obj)
        if lead_id:
            return redirect('lead_detail', pk=lead_id)
        messages.success(request, 'Follow-up logged.')
        return redirect('followup_list')
    return render(request, 'leads/followup_form.html', {'form': form, 'action': 'Log'})


def _create_recurring_followup(original):
    FollowUp.objects.create(
        lead=original.lead, deal=original.deal,
        followup_type=original.followup_type,
        subject=original.subject, notes=original.notes,
        next_followup_date=original.next_followup_date + timedelta(days=original.recurrence_days),
        is_recurring=True, recurrence_days=original.recurrence_days,
        created_by=original.created_by,
    )


@login_required
@require_POST
def followup_complete(request, pk):
    fu = get_object_or_404(FollowUp, pk=pk)
    _log_activity(lead=fu.lead, deal=fu.deal, activity_type='followup',
                  title=f'Follow-up completed: {fu.subject}', user=request.user)
    if fu.is_recurring and fu.next_followup_date:
        _create_recurring_followup(fu)
        messages.success(request, 'Follow-up completed. Next recurring follow-up created.')
    else:
        messages.success(request, 'Follow-up completed.')
    lead_pk = fu.lead_id
    fu.delete()
    if lead_pk:
        return redirect('lead_detail', pk=lead_pk)
    return redirect('followup_list')


@login_required
def followup_delete(request, pk):
    obj = get_object_or_404(FollowUp, pk=pk)
    lead_pk = obj.lead_id
    obj.delete()
    messages.success(request, 'Follow-up deleted.')
    if lead_pk:
        return redirect('lead_detail', pk=lead_pk)
    return redirect('followup_list')


# ── Import/Export CSV + XLSX (Feature 9) ───────────────────────

EXPORT_HEADERS = ['Name', 'Email', 'Phone', 'Company', 'Website', 'Lead Source', 'Type', 'Lead Owner', 'Converted', 'Created']

try:
    from openpyxl import Workbook as _XlWorkbook, load_workbook as _xl_load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

_COLUMN_MAP = {
    'name': 'name', 'lead name': 'name', 'contact name': 'name', 'full name': 'name', 'contact': 'name',
    'email': 'email', 'lead email': 'email', 'contact email': 'email', 'e-mail': 'email',
    'phone': 'phone', 'lead mobile': 'phone', 'mobile': 'phone', 'lead phone': 'phone', 'phone number': 'phone', 'contact phone': 'phone', 'tel': 'phone',
    'company': 'company', 'organization': 'company', 'org': 'company', 'company name': 'company',
    'website': 'website', 'url': 'website', 'web': 'website', 'lead website': 'website',
    'lead source': 'lead_source', 'source': 'lead_source',
    'type': 'contact_type', 'lead type': 'contact_type', 'contact type': 'contact_type',
    'lead owner': 'lead_owner', 'owner': 'lead_owner', 'assigned to': 'lead_owner',
    'converted': 'is_converted', 'status': 'is_converted',
}


def _map_row(raw_row):
    mapped = {}
    for raw_key, raw_val in raw_row.items():
        key = str(raw_key).strip().lower() if raw_key else ''
        norm = _COLUMN_MAP.get(key)
        if norm:
            mapped[norm] = str(raw_val).strip() if raw_val else ''
    return mapped


@login_required
def lead_export(request):
    fmt = request.GET.get('format', 'csv')
    leads = LeadContact.objects.all()

    if fmt == 'xlsx':
        if not HAS_OPENPYXL:
            messages.error(request, 'XLSX export is not available. Install openpyxl: pip install openpyxl')
            return redirect('lead_list')
        wb = _XlWorkbook()
        ws = wb.active
        ws.title = 'Leads'
        ws.append(EXPORT_HEADERS)
        for lead in leads:
            ws.append([
                lead.name, lead.email, lead.phone, lead.company, lead.website,
                lead.lead_source, lead.contact_type,
                lead.lead_owner.get_full_name() if lead.lead_owner else '',
                'Yes' if lead.is_converted else 'No',
                lead.created_at.strftime('%Y-%m-%d'),
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="leads_export.xlsx"'
        wb.save(response)
        return response

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="leads_export.csv"'
    writer = csv.writer(response)
    writer.writerow(EXPORT_HEADERS)
    for lead in leads:
        writer.writerow([
            lead.name, lead.email, lead.phone, lead.company, lead.website,
            lead.lead_source, lead.contact_type,
            lead.lead_owner.get_full_name() if lead.lead_owner else '',
            'Yes' if lead.is_converted else 'No',
            lead.created_at.strftime('%Y-%m-%d'),
        ])
    return response


def _parse_rows_from_file(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith('.xlsx'):
        if not HAS_OPENPYXL:
            raise ValueError('XLSX files require openpyxl. Install it with: pip install openpyxl')
        wb = _xl_load_workbook(uploaded_file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h else '' for h in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:]]
    else:
        decoded = uploaded_file.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(decoded))
        return list(reader)


@login_required
def lead_import(request):
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded = request.FILES['file']
        try:
            rows = _parse_rows_from_file(uploaded)
            imported = 0
            skipped = 0
            for row in rows:
                m = _map_row(row)
                name = m.get('name', '').strip()
                phone = m.get('phone', '').strip()
                if not name:
                    skipped += 1
                    continue
                owner = None
                owner_name = m.get('lead_owner', '').strip()
                if owner_name and owner_name != '--':
                    owner = User.objects.filter(Q(username=owner_name) | Q(first_name__icontains=owner_name)).first()
                source = m.get('lead_source', 'other').strip().lower().replace(' ', '_') or 'other'
                ctype = m.get('contact_type', 'lead').strip().lower() or 'lead'
                if ctype not in ('lead', 'deal'):
                    ctype = 'lead'
                converted_raw = m.get('is_converted', '').strip().lower()
                is_converted = converted_raw in ('yes', 'true', '1', 'converted', 'client')
                lead = LeadContact.objects.create(
                    name=name, email=m.get('email', ''),
                    phone=phone or '—',
                    company=m.get('company', ''), website=m.get('website', ''),
                    lead_source=source, contact_type=ctype,
                    lead_owner=owner, added_by=request.user,
                    is_converted=is_converted,
                )
                _log_activity(lead=lead, activity_type='note', title=f'Imported via {uploaded.name}', user=request.user)
                imported += 1
            messages.success(request, f'Imported {imported} leads. Skipped {skipped}.')
        except Exception as e:
            messages.error(request, f'Import error: {e}')
        return redirect('lead_list')
    return render(request, 'leads/lead_import.html')
